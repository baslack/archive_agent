"""
Uses an Excel file to compress and sort jobs into CD folders for burning

Logic Flow:
I. Get a list of jobs from the excel file

II. With that List do the following:

    A. Goto the folder for that job.

    B. Empty the Trash folder

    C. Copy that job folder to a working folder.

    D. ZIP the contents of that copied folder with the OSX Archive Utility

    E. Sort the ZIP into a Disk Folder
        1. Disk folder exists?
            i. Yes.
                a. Disk folder is full? Goto Next Folder
                b. Next folder doesn't exist? Create it and put ZIP in folder. Note folder.
                c. Next folder exists? Is full? Next Folder. Else, place ZIP, note folder.
            ii. No.
                a. Create Folder.
                b. Place ZIP.
                c. Note Folder.

    F. Delete the Working Copy

    G. Delete the Contents of the Job Folder

    H. Add a folder, noting the Disk Folder where the ZIP was placed

    I. Append a Disk number to that job's entry in the Excel document


"""
import openpyxl, os, time, subprocess, shutil, re, datetime

__author__ = 'Benjamin A. Slack, iam@niamjneb.com'
__version__ = '0.0.0.1'

# kBaseJobsPath = '/Volumes/JobsA'
kBaseJobsPath = os.path.expanduser('~/tmp/JobsA')
#kBaseDisksPath = '/Volumes/SGB-TITAN/_ReadyForBackup'
kBaseDisksPath = os.path.expanduser('~/tmp/Staging')
#kWorkingPath = '/Volumes/SGB-TITAN/_ReadyForBackup'
kWorkingPath = os.path.expanduser('~/tmp/Staging')
kJobFolderPrefix = '/Jobs'
kTrashFolderPrefix = '/Trash'
kDiscFolderPrefix = 'Disc'
kFullSize = 4294967296  # 4GB
kFileName = 'test.xlsx'
kPath = os.path.expanduser('~')
kSep = '/'
#kURL = kPath + kSep + kFileName
kURL = os.path.expanduser('~/tmp/Logs/test.xlsx')
dt = datetime.date.today()


class Job:
    def __init__(self, job_number):
        self.job_number = job_number
        self.location = generate_job_url(job_number)
        self.is_archived = False
        self.size = get_size(self.location)
        self.on_server = True
        self.clean = False
        self.archive = None
        self.on_disc = None
        self.ignore = False
        self.inspect()

    def dump(self):
        try:
            shutil.rmtree(self.location)
            lg.add('Job: {0}, deleted job folder @: {1}\n'.format(self.job_number, self.location))
            self.on_server = False
            return True
        except:
            lg.add('Job: {0}, unable to delete job folder @: {1}\n'.format((self.job_number, self.location)))
            return False

    def tag(self):
        try:
            tags = []
            for this_file in self.archive.files:
                tags.append(this_file.in_disc.disc_number)
            tags.sort()
            for this_tag in tags:
                path = self.location + kSep + kDiscFolderPrefix + str(this_tag).zfill(4)
                os.makedirs(path)
                lg.add('Tagged Job: {0}, with Disc: {1}.\n'.format(self.job_number, this_tag))
            return True
        except:
            lg.add('Unable to tag Job: {0}, please investigate.\n'.format(self.job_number))
            return False


    def cleanup(self):
        # empty Trash folder
        try:
            shutil.rmtree('{0}/Trash'.format(self.location))
            lg.add('Job: {0}, deleted Trash folder.\n'.format(self.job_number))
        except:
            lg.add('Job: {0}, unable to delete Trash folder.\n'.format(self.job_number))

        # clean up Image Carrier folders
        """
        1. tokenize each file name in the image carrier folder, compile a list
        2. if a file has a single token, mark it to keep
        3. if a file has multiple tokens but none match the job number, mark to delete
        4. if a file has the job number token, mark it to keep
        5. if a file has a color token
            i. look for other files that have the same token
            ii. mark the most recent to keep
            iii. mark the remainder to delete
        6. if no files are marked to keep, mark all files to keep
        7. delete files marked for deletion

        """

        def tokenize(filename):
            strip_extension = filename.rpartition('.')[0]
            break_by_whitespace = strip_extension.split()
            break_by_underscore = []
            for this_section in break_by_whitespace:
                break_by_underscore = break_by_underscore + this_section.split('_')
            break_by_hyphen = []
            for this_section in break_by_underscore:
                break_by_hyphen = break_by_hyphen + this_section.split('-')
            tokens = break_by_hyphen
            return tokens

        # compile file list

        image_carrier_path = self.location + '/Deliverables/Image_Carriers'

        files = {}
        for this_tuple in os.walk(image_carrier_path):
            this_dir = this_tuple[0]
            if this_dir[0] != '.':  # ignore dot directories
                these_files = this_tuple[2]
                for this_file in these_files:
                    if this_file[0] != '.':  # ignore dot files and hidden stores
                        this_filepath = this_dir + '/' + this_file
                        files[this_filepath] = {}
                        files[this_filepath]['name'] = this_file
                        files[this_filepath]['modified'] = os.path.getmtime(this_filepath)
                        files[this_filepath]['tokens'] = tokenize(this_file)
                        files[this_filepath]['keep'] = True

        # trap for "c###" in the last token

        pattern = 'c\d+'
        for this_filepath in files.keys():
            is_c_token = re.search(pattern, files[this_filepath]['tokens'][-1])
            if is_c_token:
                files[this_filepath]['tokens'].pop()

        # clean out not image carriers from the deletion lists

        for this_filepath in files.keys():
            is_len = files[this_filepath]['name'].rsplit('.')[-1].lower() == 'len'
            is_tif = files[this_filepath]['name'].rsplit('.')[-1].lower() == 'tif'
            is_tiff = files[this_filepath]['name'].rsplit('.')[-1].lower() == 'tiff'
            if not (is_len or is_tif or is_tiff):
                try:
                    files.pop(this_filepath, None)
                except:
                    pass

        # check for a matching job number token, if not there don't keep

        for this_filepath in files.keys():
            files[this_filepath]['keep'] = False
            for this_token in files[this_filepath]['tokens']:
                pattern = '.*{0}.*'.format(str(self.job_number))
                job_number_in_token = re.search(pattern, this_token)
                files[this_filepath]['keep'] = files[this_filepath]['keep'] or bool(job_number_in_token)

        # accumulate color tokens

        color_tokens = {}
        for this_filepath in files.keys():
            if files[this_filepath]['keep']:
                color = files[this_filepath]['tokens'][-1]  # last token should be esko's ink color
                if not (color in color_tokens.keys()):  # color token not yet created
                    color_tokens[color] = []
                    color_tokens[color].append(this_filepath)
                else:
                    color_tokens[color].append(this_filepath)

        for this_color in color_tokens.keys():  # with each color
            if len(color_tokens[this_color]) > 1:  # if there are more than one file
                dates = []  # for sorting
                file_by_dates = {}  # reverse lookup to files
                for this_filepath in color_tokens[this_color]:  # for each filepath in a color
                    dates.append(files[this_filepath]['modified'])  # append the date to the sorting list
                    file_by_dates[
                        files[this_filepath][
                            'modified']] = this_filepath  # add the filepath to the reverse lookup table
                dates.sort()  # sort the dates
                files[file_by_dates[dates.pop()]][
                    'keep'] = True  # pop the highest from the sort stack and keep that file
                while dates:
                    files[file_by_dates[dates.pop()]]['keep'] = False  #set the others to drop

        # check for single token, do this last in case the other tests have marked the file for deletion

        for this_filepath in files.keys():
            if len(files[this_filepath]['tokens']) <= 2:
                files[this_filepath]['keep'] = True

        # using the keep field, dump the dead files

        for this_filepath in files.keys():
            if not (files[this_filepath]['keep']):
                try:
                    os.remove(this_filepath)
                    lg.add('Job: {0}, removed {1}\n'.format(self.job_number, this_filepath))
                except:
                    lg.add('Job: {0}, unable to remove {1}\n'.format(self.job_number, this_filepath))

        self.clean = True

    def inspect(self):
        my_walk = os.walk(self.location)
        path, dirs, files = my_walk.next()

        try:
            files.remove('.DS_Store')  # kill the Extreme IP mac share data
        except:
            pass

        try:
            dirs.remove('config')  # kill the hidden config dirs
        except:
            pass

        # recognize disc folders

        disc_folders = []
        pattern = '([dD]+is[ck]+)(\s*[#]?)(\d+)'
        for this_dir in dirs:
            is_disc_folder = re.search(pattern, this_dir)
            if is_disc_folder:
                disc_folders.append(int(is_disc_folder.groups()[-1]))

        if len(disc_folders) > 0:
            disc_folders.sort()

            highest_disc = disc_folders[-1]

            '''
            jobs will not be spread over discs with large differences in their number.
            therefore, discard if > threshold
            '''

            threshold = 10
            for this_disc in disc_folders:
                if highest_disc - this_disc >= threshold:
                    disc_folders.remove(this_disc)

            if len(dirs) == len(disc_folders) and len(dirs) > 0:
                #  job has been archived
                self.is_archived = True
                self.on_disc = disc_folders
                lg.add('Job: {0}, has already been archived on {1}.\n'.format(self.job_number, repr(self.on_disc)))
            elif len(dirs) > len(disc_folders) and len(dirs) > 0:
                #  job has directories other than disc folders, archive the job
                self.is_archived = False
                self.ignore = False
                lg.add('Job: {0}, will be archived.\n'.format(self.job_number))
            else:
                # job has fewer directories than it does disc folders, this can not happen
                try:
                    raise Exception('Job inspection failed. More disc folders than directories in job.')
                except:
                    self.ignore = True
                    lg.add('Job: {0}, inspection failed. Ignoring.\n'.format(self.job_number))

        if len(dirs) == 0:
            try:
                raise Exception('Job Folder Is Empty.')  #no directories means something is wrong, get a human
            except:
                self.ignore = True
                lg.add('Job: {0}, is empty. Ignoring.\n'.format(self.job_number))
        elif len(disc_folders) == 0:
            lg.add('Job: {0}, will be archived.\n'.format(self.job_number))


class File:
    def __init__(self, url):
        self.location = url
        self.name = url.rsplit('/', 1)[1]
        self.size = get_size(url)
        self.is_placed = False
        self.in_disc = False

    def add2disc(self, disc):
        if self.size + disc.size >= kFullSize:
            return False
        else:
            shutil.move(self.location, disc.location)
            self.is_placed = True
            self.in_disc = disc
            disc.size += self.size
            if disc.size >= kFullSize:
                disc.is_full = True
            self.location = disc.location + kSep + self.name
            lg.add('File: {0}, placed @: {1}\n'.format(self.name, self.location))
            return True


class Archive:
    def __init__(self, job):
        self.job = job
        self.files = []

        # archive the job
        command = 'ditto'
        args = '-c -k --sequesterRsrc --keepParent'
        zip_name = str(job.job_number) + '.zip'
        zip_path = kWorkingPath + kSep + zip_name
        do_this = '{0} {1} {2} {3}'.format(command, args, job.location, zip_path)
        subprocess.call(do_this, shell=True)
        zip_size = get_size(zip_path)

        # break it up if required
        if zip_size >= kFullSize:
            command = 'zip'
            args = '-q -s 2g'
            split_path = kWorkingPath + kSep + '{0}_split.zip'.format(job.job_number)
            do_this = '{0} {1} {2} {3}'.format(command, args, split_path, zip_path)
            subprocess.call(do_this, shell=True)
            os.remove(zip_path)
            these_files = os.listdir(kWorkingPath)
            for this_file in these_files:
                if this_file.find('_split') > 0:  # contains the split file name
                    self.files.append(File(kWorkingPath + kSep + this_file))
                    lg.add('Archive created: {0}\n'.format(kWorkingPath + kSep + this_file))
        else:
            self.files.append(File(zip_path))
            lg.add('Archive created: {0}\n'.format(zip_path))
        job.is_archived = True
        job.archive = self


class Disc:
    def __init__(self, disc_number):
        self.disc_number = disc_number
        self.folder_name = kDiscFolderPrefix + str(self.disc_number).zfill(4)
        self.location = kBaseDisksPath + kSep + self.folder_name

        try:
            os.makedirs(self.location)
            lg.add('Created disc folder: {0}\n'.format(self.location))
        except:
            pass
        self.size = get_size(self.location)
        if self.size >= kFullSize:
            self.is_full = True
        else:
            self.is_full = False
        self.contents = os.listdir(self.location)


class Log:
    def __init__(self, path):
        self.path = path
        self.name = self.path.rsplit('/')[-1]
        try:
            self.size = get_size(self.path)
        except:
            self.size = 0
        self.file = open(self.path, 'a+', 1)


    def open(self):
        self.file = open(self.path, 'a+', 1)

    def close(self):
        self.file.close()

    def add(self, string):
        print(string)
        self.file.write(string)


class Manager:
    def __init__(self, url):
        self.excel_url = url
        self.job_list = []
        self.disc_catalog = []
        self.get_job_list(url)
        self.setup_disc_catalog()


    def get_job_list(self, url):
        """
        Setups up the master job list and attaches the Excel file log to the manager object.
        """
        try:
            self.wb = openpyxl.load_workbook(url)
        except:
            print('No excel file at {0)'.format(url))
            lg.add('Excel file: {0}, does not exist.\n'.format(url))
            return False
        ws = self.wb.active
        max_row = int(ws.max_row) + 1
        max_col = ws.max_column

        r, c, c2 = 1, 1, 5
        parsed_jobs = []
        while r < max_row:
            try:
                job_number = int(ws.cell(row=r, column=c).value)  # first entry is a job number
            except:
                r += 1
                continue
            if not (ws.cell(row=r, column=c2).value):  # disc already attached to job
                # check for double entries
                if not (job_number in parsed_jobs):
                    self.job_list.append(Job(job_number))
                    parsed_jobs.append(job_number)
                else:
                    lg.add('Duplicate Entry for Job: {0} found. Skipping Entry @ row: {1}\n'.format(job_number, r))
            r += 1
        return True

    def setup_disc_catalog(self, url=kBaseDisksPath):
        """
        searches the base disc directory for existing disc folders, sizes them and populates the disc catalog
        """

        """
        1. get the list of folders in the staging directory, don't read deeper than the root level
        2. remove any non-disc folders from that listing
        3. with the remaining folders, calculate the size of each disc, store the value
        4. if no disc folders exists, prompt for a disc number to start with, create that folder, set it's size to 0
        """

        directories = os.walk(url).next()[1]

        for this_dir in directories:  # drop the non-disc dirs
            if this_dir.find('Disc') == -1:
                directories.remove(this_dir)

        if len(directories) > 0:  # one or more disc directories

            for index in range(len(directories)):  # drop the "Disc" prefix and convert the ids to ints
                directories[index] = int(directories[index].strip('Disc'))

            for this_disc_number in directories:
                self.disc_catalog.append(Disc(this_disc_number))
        else:  # no disc directories
            while not ('disc_number' in locals()):  # ask for a disc number until you get a valid one
                try:
                    disc_number = int(input('Enter a starting Disc #: '))
                except:
                    print('Disc number is invalid, please enter an integer disc number.')
            self.disc_catalog.append(Disc(disc_number))

        return True

    def get_last_disc(self):
        disc_numbers = []
        for this_disc in self.disc_catalog:
            disc_numbers.append(this_disc.disc_number)
        disc_numbers.sort()
        return disc_numbers[-1]

    def update_workbook(self):
        ws = self.wb.active
        c, c2 = 1, 5

        for this_row in range(1, ws.max_row + 1):
            try:
                job_number = int(ws.cell(row=this_row, column=c).value)
            except:
                continue
            for this_job in self.job_list:
                if this_job.job_number == job_number:
                    disc_entry = ''
                    if this_job.on_disc:  # was already on disc
                        for this_disc in this_job.on_disc:
                            disc_entry += '{0},'.format(str(this_disc))
                    else:  # was archived
                        for this_file in this_job.archive.files:
                            disc_entry += '{0},'.format(str(this_file.in_disc.disc_number))
                    ws.cell(row=this_row, column=c2).value = disc_entry[0:-1]
        self.wb.save(self.excel_url)


def generate_job_url(job):
    """

    :param job: the
     job number to generate the URL for
    :return: the string of the URL to the job on the server
    """

    if type(job) != type(''):
        job = str(job)
    folderNumber = job[len(job) - 1:len(job)]  # get the last number of the job
    return kBaseJobsPath + kJobFolderPrefix + folderNumber + kSep + job


def generate_working_url(job):
    """

    :param job: the job number to generate a working folder for
    :return: the string of the URL of that working folder
    """
    if type(job) != type(''):
        job = str(job)
    return kWorkingPath + kSep + job


def generate_disc_url(disc):
    """

    :param disc: the disc number to generate path to string for
    :return: the path to string
    """
    if not isinstance(disc, str):
        disc = str(disc).zfill(4)
    return kBaseDisksPath + kDiscFolderPrefix + disc


def get_size(path):
    total_size = 0
    try:
        os.walk(path).next()
        for dir_path, dir_names, filenames in os.walk(path):
            for f in filenames:
                fp = os.path.join(dir_path, f)
                total_size += os.path.getsize(fp)
    except:
        total_size = os.path.getsize(path)
    return total_size


lg = Log(kWorkingPath + kSep + 'Log_{0}-{1}-{2}.txt'.format(dt.month, dt.day, dt.year))

if __name__ == "__main__":
    mngr = Manager(kURL)

    # code for bucketing the jobs
    for this_job in mngr.job_list:
        if not this_job.ignore and not this_job.is_archived:
            this_job.cleanup()
            this_job.archive = Archive(this_job)
            for this_file in this_job.archive.files:
                index = 0
                while index < len(mngr.disc_catalog):
                    try:
                        if not (this_file.add2disc(mngr.disc_catalog[index])):
                            index += 1
                        else:
                            break
                    except:
                        pass
                if not this_file.is_placed:
                    new_disc_number = mngr.get_last_disc() + 1
                    new_disc = Disc(new_disc_number)
                    mngr.disc_catalog.append(new_disc)
                    this_file.add2disc(new_disc)
        lg.file.flush()

    # code for cleaning up the directory and adding disk tags
    accepted_inputs = ['y', 'n', 'a']
    selection = None
    while not (selection in accepted_inputs):
        selection = raw_input('Proceed with job dumps and folder tagging? (y/n/a):')
        selection = str(selection).lower()
        if not (selection in accepted_inputs):
            print('Selection Invalid, please select (y/n/a).\n')

    if selection != 'n':
        for this_job in mngr.job_list:
            if not (this_job.on_disc) and not (this_job.ignore):
                if selection == 'y':
                    confirm = None
                    confirm_inputs = ['y', 'n']
                    while not (confirm in confirm_inputs):
                        confirm = raw_input('Delete Job: {0} @ {1}? (y/n): '.format(this_job.job_number, this_job.location))
                        confirm = str(confirm).lower()
                        if not (confirm in confirm_inputs):
                            print('Selection not recognized. Please enter (y/n).\n')
                    if confirm == 'y':
                        this_job.dump()
                        this_job.tag()
                    else:
                        continue
                elif selection == 'a':
                    this_job.dump()
                    this_job.tag()
    else:
        lg.add('Leaving jobs on server.\n')

    # code for updating the excel doc
    mngr.update_workbook()

lg.close()
