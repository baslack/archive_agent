"""
Uses an Excel file to compress and sort jobs into CD folders for burning

Logic Flow:
I. Get a list of jobs from the excel file

II. With that List do the following:

    A. Goto the folder for that job.

    B. Empty the Trash folder

    C. Copy that job folder to a working folder.

    D. ZIP the contents of that copied folder with the OSX Archive Utility

    E. Sort the ZIP into a Disk Folder
        1. Disk folder exists?
            i. Yes.
                a. Disk folder is full? Goto Next Folder
                b. Next folder doesn't exist? Create it and put ZIP in folder. Note folder.
                c. Next folder exists? Is full? Next Folder. Else, place ZIP, note folder.
            ii. No.
                a. Create Folder.
                b. Place ZIP.
                c. Note Folder.

    F. Delete the Working Copy

    G. Delete the Contents of the Job Folder

    H. Add a folder, noting the Disk Folder where the ZIP was placed

    I. Append a Disk number to that job's entry in the Excel document


"""
import openpyxl, os, time, subprocess, shutil

__author__ = 'Benjamin A. Slack, iam@niamjneb.com'
__version__ = '0.0.0.1'

kBaseJobsPath = '/Volumes/JobsA'
kBaseDisksPath = '/Volumes/SGB-TITAN/_ReadyForBackup'
kWorkingPath = '/Volumes/SGB-TITAN/_ReadyForBackup'
kJobFolderPrefix = '/Jobs'
kTrashFolderPrefix = '/Trash'
kDiscFolderPrefix = '/Disc'
kFullSize = '4294967296'  # 4GB
kFileName = 'test.xlsx'
kPath = os.path.expanduser('~')
kSep = '/'
kURL = kPath + kSep + kFileName

disc_catalog = {}
job_list = []
simulate = True
log_buffer = []


def dump_log(filename='log_'+str(int(time.time()))):
    '''

    :param filename: name of the log file to open
    :return: file object for the log file
    '''
    if not simulate:
        path = os.path.abspath('.')
    else:
        path = kPath + '/tmp'
    url = path + kSep + filename
    log = open(url, 'w+')
    log.writelines(log_buffer)
    log.close()
    return url

def get_list(url):
    """

    :param url: The location of the excel file we're using
    :return: a list of dictionaries, containing 'id', 'row' and 'col' of the job in the excel file
    """

    wb = openpyxl.load_workbook(url)
    ws = wb.active
    r, c = 1, 1
    while ws.cell(row=r, column=c).value:
        item = {}
        try:
            item['id'] = int(ws.cell(row=r, column=c).value)
            item['row'] = r
            item['col'] = c
        except:
            r += 1
            continue
        job_list.append(item)
        r += 1
    return 0


def generate_job_url(job):
    """

    :param job: the
     job number to generate the URL for
    :return: the string of the URL to the job on the server
    """

    if type(job) != type(''):
        job = str(job)
    folderNumber = job[len(job) - 1:len(job)]  # get the last number of the job
    return kBaseJobsPath + kJobFolderPrefix + folderNumber + kSep + job


def generate_working_url(job):
    """

    :param job: the job number to generate a working folder for
    :return: the string of the URL of that working folder
    """
    if type(job) != type(''):
        job = str(job)
    return kWorkingPath + kSep + job


def generate_disc_url(disc):
    """

    :param disc: the disc number to generate path to string for
    :return: the path to string
    """
    if not isinstance(disc, str):
        disc = str(disc).zfill(4)
    return kBaseDisksPath + kDiscFolderPrefix + disc


def dump_trash(job):
    """

    :param job: The job number to dump the trash of
    :return:
    """
    deleteMe = generate_job_url(job) + kTrashFolderPrefix
    if not simulate:
        shutil.rmtree(deleteMe)
    else:
        print(deleteMe)
    return 0


def clean_IC(job):
    """

    :param job: job number to clean up the image carrier folder for
    :return: 0 if successful
    """

    """
    1. tokenize each file name in the image carrier folder, compile a list
    2. if a file has a single token, mark it to keep
    3. if a file has multiple tokens but none match the job number, mark to delete
    4. if a file has the job number token, mark it to keep
    5. if a file has a color token
        i. look for other files that have the same token
        ii. mark the most recent to keep
        iii. mark the remainder to delete
    6. if no files are marked to keep, mark all files to keep
    7. delete files marked for deletion

    """

    def tokenize(filename):
        strip_extension = filename.rpartition('.')[0]
        break_by_whitespace = strip_extension.split()
        break_by_underscore = []
        for this_section in break_by_whitespace:
            break_by_underscore = break_by_underscore + this_section.split('_')
        break_by_hyphen = []
        for this_section in break_by_underscore:
            break_by_hyphen = break_by_hyphen + this_section.split('-')
        tokens = break_by_hyphen
        return tokens

    # compile file list

    image_carrier_path = kPath+'/tmp/Files'
    #image_carrier_path = generate_job_url(job) + '/Deliverables/Image_Carriers'

    files = {}
    for this_tuple in os.walk(image_carrier_path):
        this_dir = this_tuple[0]
        if this_dir[0] != '.':  # ignore dot directories
            these_files = this_tuple[2]
            for this_file in these_files:
                if this_file[0] != '.':  # ignore dot files and hidden stores
                    this_filepath = this_dir + '/' + this_file
                    files[this_filepath] = {}
                    files[this_filepath]['name'] = this_file
                    files[this_filepath]['modified'] = os.path.getmtime(this_filepath)
                    files[this_filepath]['tokens'] = tokenize(this_file)
                    files[this_filepath]['keep'] = True

    # check for a matching job number token, if not there don't keep

    for this_filepath in files.keys():
        if not(str(job) in files[this_filepath]['tokens']):
            files[this_filepath]['keep'] = False

    # accumulate color tokens

    color_tokens = {}
    for this_filepath in files.keys():
        if files[this_filepath]['keep']:
            color = files[this_filepath]['tokens'][-1] # last token should be esko's ink color
            if color_tokens.get(color, True): # color token not yet created
                color_tokens[color] = []
                color_tokens[color].append(this_filepath)
            else:
                color_tokens[color].append(this_filepath)

    for this_color in color_tokens.keys():  # with each color
        if len(color_tokens[this_color]) > 1:  # if there are more than one file
            dates = []  # for sorting
            file_by_dates = {}  # reverse lookup to files
            for this_filepath in color_tokens[this_color]:  # for each filepath in a color
                dates.append(files[this_filepath]['modified'])  # append the date to the sorting list
                file_by_dates[files[this_filepath]['modified']] = this_filepath  # add the filepath to the reverse lookup table
            dates.sort()  # sort the dates
            files[file_by_dates[dates.pop()]]['keep'] = True  # pop the highest from the sort stack and keep that file
            while dates:
                files[file_by_dates[dates.pop()]]['keep'] = False  #set the others to drop

    # check for single token, do this last in case the other tests have marked the file for deletion

    for this_filepath in files.keys():
        if len(files[this_filepath]['tokens']) == 1:
            files[this_filepath]['keep'] = True

    # using the keep field, dump the dead files

    for this_filepath in files.keys():
        if not(files[this_filepath]['keep']):
            if not simulate:
                try:
                    os.remove(this_filepath)
                    log_buffer.append('Deleted: {0}\n'.format(this_filepath))
                except:
                    log_buffer.append(('Could not delete: {0}\n'.format(this_filepath)))
            else:
                log_buffer.append('Will delete: {0}\n'.format(this_filepath))

    '''
    for this_filepath in files.keys():
        print('path: {0}, settings: {1}'.format(this_filepath, repr(files[this_filepath])))
    '''

    return 0



def copy_job(job):
    """

    :param job: The job number to copy to the work folder
    :return: the copy's url
    """
    copyMeFrom = generate_job_url(job)
    copyMeTo = generate_working_url(job)
    command = 'ditto'
    do_this = [command, copyMeFrom, copyMeTo]
    if not simulate:
        try:
            subprocess.call(do_this)
            log_buffer.append("Copied: {0} to {1}\n".format(copyMeFrom, copyMeTo))
        except:
            log_buffer.append('Unable to copy: {0} to {1}\n'.format(copyMeFrom, copyMeTo))
    else:
        log_buffer.append('Will copy: {0} to {1}\n'.format(copyMeFrom, copyMeTo))
    return copyMeTo


def zip_job(job):
    """

    :param job: The job number to zip using Archive Utility
    :return: zip file path
    """

    command = ['ditto']
    args = ['-c', '-k', '--sequesterRsrc', '--keepParent']
    working_path = [generate_working_url(job)]
    zip_path = [kWorkingPath + '/' + str(job) + '.zip']
    do_this = command + args + working_path + zip_path

    if not simulate:
        try:
            subprocess.call(do_this)
            log_buffer.append('Zipped: {0} to {1}\n'.format(job, zip_path[0]))
        except:
            log_buffer.append('Unable to zip: {0} to {1}\n'.format(job, zip_path[0]))
    else:
            log_buffer.append('Will zip: {0} to {1}\n'.format(job, zip_path[0]))
    return zip_path[0]


def bucket_job(zip):
    """

    :param zip: zip to put into disc folder
    :return: disc number ZIP was put in

    1. check zip size
    2. check for disc folders
        a. found
            i. for each
                1. check size
                2. check fit
                    a. fits
                        i. move into folder
                    b. doesn't fit
                        i. next folder
                3. no fits
                    a. create folder
                    b. move into folder

        b. not found
            i. create folder
            ii. copy into folder

    """


def inspect_discs():
    """
    searches the staging directory for existing disc folders, sizes them and populates the disc catalog
    :return:
    """

    """
    1. get the list of folders in the staging directory, don't read deeper than the root level
    2. remove any non-disc folders from that listing
    3. with the remaining folders, calculate the size of each disc, store the value
    4. if no disc folders exists, prompt for a disc number to start with, create that folder, set it's size to 0
    """
    #staging = kBaseDisksPath
    staging = kPath + '/tmp/Staging2'

    directories = os.walk(staging).next()[1]

    for this_dir in directories: # drop the non-disc dirs
        if this_dir.find('Disc') == -1:
            directories.remove(this_dir)

    if len(directories) > 0:  # one or more disc directories

        for index in range(len(directories)): # drop the "Disc" prefix and convert the ids to ints
            directories[index] = int(directories[index].strip('Disc'))

        for this_dir in directories:
            disc_catalog[this_dir] = {}
            disc_catalog[this_dir]['path'] = staging + kDiscFolderPrefix + str(this_dir).zfill(4)
            disc_catalog[this_dir]['size'] = get_size(disc_catalog[this_dir]['path'])
    else:  # no disc directories
        while not('disc' in locals()):  # ask for a disc number until you get a valid one
            try:
                disc = int(input('Enter a starting Disc #: '))
            except ValueError, NameError:
                print('Disc number is invalid, please enter an integer disc number.')

        os.mkdir(staging + kDiscFolderPrefix + str(disc).zfill(4))  # create a new disc directory
        log_buffer.append('Created Directory: {0}\n'.format(staging + kDiscFolderPrefix + str(disc).zfill(4)))
        disc_catalog[disc] = {}  # add an entry to the catalog
        disc_catalog[disc]['path'] = staging + kDiscFolderPrefix + str(disc).zfill(4)
        disc_catalog[disc]['size'] = 0

    return disc_catalog


def get_size(path = '.'):
    total_size = 0
    for dir_path, dir_names, filenames in os.walk(path):
        for f in filenames:
            fp = os.path.join(dir_path, f)
            total_size += os.path.getsize(fp)
    return total_size


def check_disc(disc):
    """
    :param disc: the disc number of the folder to check for size
    :return: full or not
    """

    try:
        if disc_catalog[disc]['size'] >= kFullSize:
            print('Disc {0} is full.'.format(disc))
            return True
        else:
            print('Disc {0} has room.'.format(disc))
            return False
    except KeyError:
        print('Disc {0} Does Not Exist.'.format(disc))
        return -1


def create_disc():
    """
    :return the number of the new disc
    """
    discs = disc_catalog.keys()
    discs.sort()
    last_disc = discs.pop()
    new_disc = last_disc + 1
    if not simulate:
        try:
            os.mkdir(generate_disc_url(new_disc))
            disc_catalog[new_disc]['path'] = generate_disc_url(new_disc)
            disc_catalog[new_disc]['size'] = 0
            log_buffer.append('Created new disc foolder: {0} at {1}\n'.format(new_disc, generate_disc_url(new_disc)))
        except:
            log_buffer.append('Unable to create new disc folder: {0} at {1}\n'.format(new_disc, generate_disc_url(new_disc)))
    else:
        log_buffer.append('Will create new disc folder: {0} at {1}\n'.format(new_disc, generate_disc_url(new_disc)))
    return new_disc


def dump_job(job):
    """

    :param job: the job number to dump the contents on
    :return:
    """
    if not simulate:
        try:
            shutil.rmtree(generate_job_url(job))
            log_buffer.append('Dumped job#: {0} from {1}\n'.format(job, generate_job_url(job)))
        except:
            log_buffer.append('Unable to dump job#: {0} from {1}\n'.format(job, generate_job_url(job)))
    else:
        log_buffer.append('Will dump job#: {0} from {1}\n'.format(job, generate_job_url(job)))
    return 0


def tag_job(job, disc):
    """

    :param job: the job number to add the disc tag folder to
    :param disc: the disc number for the tag
    :return:
    """
    if not simulate:
        try:
            os.mkdir(generate_job_url(job)+kSep+str(disc))
            log_buffer.append('Tagged Job# {0} with Disc# {1}\n'.format(job, disc))
        except:
            log_buffer.append('Unable to tag Job# {0} with Disc# {1}\n'.format(job, disc))
    else:
        log_buffer.append('Will tag job# {0} with Disc# {1}\n'.format(job, disc))
    return 0


def compile_excel_tags(job, disc):
    '''

    :param job_list: the list of dictionaries gathered by get_list
    :param job: a job number
    :param disc:  a disc number
    :return: the modified listing
    '''

    for this_item in job_list:
        if this_item['id'] == job:
            this_item['disc'] = disc
    return 0

def dump_list_to_excel(url):
    '''

    :param url: path of the excel file to update
    :return:
    '''
    wb = openpyxl.load_workbook(url)
    ws = wb.active
    for this_item in job_list:
        if not simulate:
            try:
                ws.cell(row=this_item['row'], column=4).value = this_item['disc']
                log_buffer.append('Updated Excel File {0}, job# {1} to disc# {2}\n'.format(url, this_item['id'], this_item['disc']))
            except:
                log_buffer.append('Problem updating Excel File {0}, job# {1} to disc# {2}\n'.format(url, this_item['id'], this_item['disc']))
    wb.save(url)
    return 0


if __name__ == "__main__":
    get_list(kURL)
    print(repr(job_list))
    print(generate_job_url(17545))
    print(generate_working_url(17545))
    print(generate_disc_url(1534))
    print(dump_trash(15687))
    print(copy_job(68975))
    print(zip_job(566843))
    print(clean_IC(4044906))
    print(repr(inspect_discs()))
    print(check_disc(1004))
    print(check_disc(1005))
    print(repr(disc_catalog))
    dump_log()